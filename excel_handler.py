"""
Export reconciliation results to the claims Excel format, and provide the
naming conventions for the statement file, the Excel file, and renamed
receipts.

Columns: Posting Date | Transaction Date | No. | PDF/Reference | Description |
         Currency | Foreign Amount | Amount (RM) | Category | Claim Notes |
         Supporting Doc Obtained?

"PDF/Reference" is plain text - the renamed receipt filename, for reference -
not a clickable link. (An earlier version tried various link mechanisms;
they proved too fragile across extraction locations/renaming/Excel versions,
so this now just states the filename plainly.)
"""
import io
import re
import zipfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CLAIMABLE_FILL = PatternFill("solid", start_color="D9EBEF", end_color="D9EBEF")
NON_CLAIMABLE_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HEADER_FILL = PatternFill("solid", start_color="1F4E5F", end_color="1F4E5F")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
NEGATIVE_FONT = Font(name="Arial", size=10, color="C00000")
TITLE_FONT = Font(name="Arial", bold=True, size=13, color="1F4E5F")
THIN_BORDER = Border(*(Side(style="thin", color="D0D0D0"),) * 4)

COLUMNS = [
    "Posting Date",
    "Transaction Date",
    "No.",
    "Reference",
    "Description",
    "Currency",
    "Foreign Amount",
    "Amount (RM)",
    "Category",
    "Claim Notes",
    "Supporting Doc Obtained?",
]


# ── Naming helpers ─────────────────────────────────────────────────────────────

def sanitize(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "", s or "")
    return s or "Unknown"


def month_number(month_name: str) -> int:
    """Best-effort month name -> 1-12. Falls back to the current month."""
    name = (month_name or "").strip()
    for fmt in ("%B", "%b"):
        try:
            return datetime.strptime(name, fmt).month
        except ValueError:
            continue
    return datetime.now().month


def normalize_ddmmyy(date_str: str, fallback_posting_date: str = "", year_hint: int | None = None) -> str:
    """Return a 6-digit DDMMYY string from whatever date info is available."""
    year_hint = year_hint or datetime.now().year
    s = (date_str or "").strip()

    if re.fullmatch(r"\d{6}", s):
        return s
    if re.fullmatch(r"\d{8}", s):  # legacy DDMMYYYY -> DDMMYY
        return s[:4] + s[6:8]

    m = re.search(r"(\d{1,2})[/\-\s]+([A-Za-z]+|\d{1,2})[/\-\s]+(\d{2,4})?", s)
    if not m:
        m = re.search(r"(\d{1,2})\s+([A-Za-z]+)", fallback_posting_date or "")
        if m:
            day, mon_raw, yr_raw = m.group(1), m.group(2), None
        else:
            return datetime.now().strftime("%d%m%y")
    else:
        day, mon_raw, yr_raw = m.group(1), m.group(2), m.group(3)

    mon = int(mon_raw) if mon_raw.isdigit() else month_number(mon_raw)
    yr = int(yr_raw) if yr_raw else year_hint
    return f"{int(day):02d}{mon:02d}{yr % 100:02d}"


def zero_pad_day(date_str: str) -> str:
    """'8 May' -> '08 May'. Leaves anything that doesn't start with a day number unchanged."""
    m = re.match(r"^\s*(\d{1,2})(\s+.*)$", date_str or "")
    if not m:
        return date_str or ""
    return f"{int(m.group(1)):02d}{m.group(2)}"


def statement_filename(month: str, start: str, end: str) -> str:
    return f"{month} Statement ({start} - {end}).pdf"


def excel_filename(month: str, year: int) -> str:
    return f"{month_number(month):02d}. {month} Claimable Invoice ({year}).xlsx"


def sanitize_folder_name(s: str) -> str:
    """Strip characters that are invalid in file/folder names, but keep spaces."""
    return re.sub(r'[\\/:*?"<>|]+', "", s or "").strip() or "Unknown Bank"


def receipt_filename(t: dict, ext: str, statement_year: int) -> str:
    r = t.get("matched_receipt")
    if r and r.get("is_myipo") and r.get("application_number"):
        return f"{sanitize(r['application_number'])}.{ext}"
    date_str = normalize_ddmmyy(
        r.get("date", "") if r else "", fallback_posting_date=t.get("posting_date", ""), year_hint=statement_year
    )
    doc_type = r.get("document_type", "Receipt") if r else "Receipt"
    company = sanitize(r.get("company_name", "")) if r else "Unknown"
    if t.get("foreign_amount"):
        amount = r.get("amount") if r else t["foreign_amount"]
        currency_prefix = (t.get("currency") or (r.get("currency") if r else "") or "USD").upper()
    else:
        amount = t["amount_rm"]
        currency_prefix = "RM"
    return f"{t['row_no']}. {date_str}_{doc_type}_{company}_{currency_prefix}{abs(amount):.2f}.{ext}"


def myipo_reference_label(t: dict, statement_year: int) -> str:
    """Synthetic Reference text (no actual file) for a MyIPO transaction recognised from its
    description that has no uploaded/matched receipt - follows the same naming convention as a
    real renamed receipt, using 'MyIPO' as the placeholder merchant name."""
    date_str = normalize_ddmmyy("", fallback_posting_date=t.get("posting_date", ""), year_hint=statement_year)
    return f"{t['row_no']}. {date_str}_Receipt_MyIPO_RM{abs(t['amount_rm']):.2f}"


# ── Excel builder ──────────────────────────────────────────────────────────────

def build_excel(transactions: list[dict], bank: str = "", month: str = "", year: int = 0,
                 start_date: str = "", end_date: str = "") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Claims"

    n_cols = len(COLUMNS)

    # ── Title row: "{Bank} ({Start} - {End})", merged and centered above the table ──
    if bank or start_date or end_date:
        title_text = bank
        if start_date or end_date:
            title_text += f" ({zero_pad_day(start_date)} - {zero_pad_day(end_date)})"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = ws.cell(row=1, column=1, value=title_text.strip())
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        header_row = 2
    else:
        header_row = 1

    for c, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=c, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 32

    data_start_row = header_row + 1
    for i, t in enumerate(transactions, start=data_start_row):
        fill = CLAIMABLE_FILL if t.get("claimable") else NON_CLAIMABLE_FILL
        r = t.get("matched_receipt")
        has_doc = bool(r) and bool(t.get("renamed_filename"))
        foreign_amt = round(t["foreign_amount"], 2) if t.get("foreign_amount") else None
        rm_amt = round(t["amount_rm"], 2)

        pdf_ref_value = ""
        if has_doc:
            pdf_ref_value = t["renamed_filename"]
        elif t.get("category") == "MyIPO":
            pdf_ref_value = myipo_reference_label(t, year)

        if has_doc and r.get("is_myipo"):
            doc_status = "MyIPO"
        elif has_doc:
            doc_status = "Yes"
        else:
            doc_status = "No"

        row_values = [
            t.get("posting_date", ""),
            t.get("transaction_date", ""),
            t.get("row_no", i - data_start_row + 1),
            pdf_ref_value,
            t.get("description", ""),
            t.get("currency", "MYR"),
            foreign_amt,
            rm_amt,
            t.get("category", ""),
            t.get("claim_notes", ""),
            doc_status,
        ]
        for c, val in enumerate(row_values, start=1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if c == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if c in (7, 8):
                cell.number_format = "0.00"
                if (c == 7 and foreign_amt is not None and foreign_amt < 0) or (c == 8 and rm_amt < 0):
                    cell.font = NEGATIVE_FONT

    # Column widths: PDF/Reference wide enough for a full renamed filename without
    # squeezing; Supporting Doc Obtained? narrower, relying on wrap_text for its header.
    widths = [13, 14, 6, 55, 40, 9, 13, 12, 20, 32, 14]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = f"A{data_start_row}"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ── Zip packaging ──────────────────────────────────────────────────────────────

def build_zip(bank: str, month: str, year: int, start: str, end: str,
              statement_bytes: bytes, excel_bytes: bytes, transactions: list[dict]) -> bytes:
    """
    Root of the ZIP (so that Windows' default 'Extract All', which creates a folder
    named after the ZIP itself, produces exactly this layout with no extra nesting):

      {Month}_Statement/
        {Month} Statement (Start - End).pdf
        {MonthNumber}. {Month} Claimable Invoice ({Year}).xlsx
      Invoice Reference/
        <matched receipts, renamed>
    """
    bio = io.BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(f"{month}_Statement/{statement_filename(month, start, end)}", statement_bytes)
        z.writestr(f"{month}_Statement/{excel_filename(month, year)}", excel_bytes)
        for t in transactions:
            r = t.get("matched_receipt")
            if r and t.get("renamed_filename"):
                z.writestr(f"Invoice Reference/{t['renamed_filename']}", r["content"])
    return bio.getvalue()